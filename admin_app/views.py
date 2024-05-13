from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, HttpResponseNotFound
from django.urls import reverse
from django.views import View
from django.http import HttpResponse
from django.db.models import Q
from django.utils.translation import gettext_lazy as _
from django.contrib import messages
import decimal
from django.http import FileResponse
from django.shortcuts import get_object_or_404
import os, json
from django.http import JsonResponse

from django.conf import settings

from django.template.loader import render_to_string
from django.core.mail import EmailMessage

from manager import models as ManagerModels
from clinicMas import models as ClinicModels
from manager import forms as ManagerForms
from admin_app.mixins import AdminAndAuthenticatedAccessMixin

import random, os
from openpyxl import load_workbook, Workbook
from io import BytesIO
from openpyxl.styles import Font
from openpyxl.drawing.image import Image

from django.conf import settings


class HomeView(AdminAndAuthenticatedAccessMixin, View):
    template_name = "admin_app/index.html"
    template_partial_home = "admin_app/partials/__home.html"

    template_partial_scheme_list = "admin_app/partials/__schemes_list.html"
    context_data = {}

    def get(self, request):
        total_credit = 0
        total_debit = 0
        schemes = ManagerModels.Scheme.objects.filter(status=True).order_by("-id")
        transactions = ManagerModels.Transaction.objects.all()[:10]

        for scheme in schemes:
            total_credit = total_credit + scheme.credit

        for transaction in transactions:
            total_debit = total_debit + transaction.amount_used

        self.context_data["total_credit"] = total_credit
        self.context_data["total_debit"] = total_debit
        self.context_data["schemes"] = schemes
        self.context_data["transactions"] = transactions
        self.context_data["beneficiaries"] = ManagerModels.FamilyMember.objects.filter(status=True)

        return render(request, template_name=self.template_name, context=self.context_data)

    def post(self, request):
        keyword = request.POST.get("search-keyword")
        if not keyword:
            return render(request, template_name=self.template_partial_home, context=self.context_data)

        schemes = ManagerModels.Scheme.objects.filter(Q(name__icontains=keyword) | Q(insurance_number__icontains=keyword))
        self.context_data["schemes"] = schemes
        return render(request, template_name=self.template_partial_scheme_list, context=self.context_data)


class SchemeListView(AdminAndAuthenticatedAccessMixin, View):
    template_name = "admin_app/schemes.html"
    template_partial_home = "admin_app/partials/__home.html"

    template_partial_scheme_list = "admin_app/partials/__schemes_list.html"
    context_data = {}

    def get(self, request):
        total_credit = 0
        total_debit = 0
        schemes = ManagerModels.Scheme.objects.filter(status=True).order_by("-id")
        transactions = ManagerModels.Transaction.objects.all()[:10]

        for scheme in schemes:
            total_credit = total_credit + scheme.credit

        for transaction in transactions:
            total_debit = total_debit + transaction.amount_used

        self.context_data["total_credit"] = total_credit
        self.context_data["total_debit"] = total_debit
        self.context_data["schemes"] = schemes
        self.context_data["transactions"] = schemes
        self.context_data["beneficiaries"] = ManagerModels.FamilyMember.objects.filter(status=True)

        return render(request, template_name=self.template_name, context=self.context_data)

    def post(self, request):
        keyword = request.POST.get("search-keyword")
        if not keyword:
            return render(request, template_name=self.template_partial_home, context=self.context_data)

        schemes = ManagerModels.Scheme.objects.filter(Q(name__icontains=keyword) | Q(insurance_number__icontains=keyword))
        self.context_data["schemes"] = schemes
        print("schemes: ", schemes)
        return render(request, template_name=self.template_partial_scheme_list, context=self.context_data)


class SchemeDeleteView(AdminAndAuthenticatedAccessMixin, View):
    def get(self, request, insurance_number):
        scheme = get_object_or_404(ManagerModels.Scheme, insurance_number=insurance_number)
        scheme.delete()
        return redirect(reverse("admin_app:schemes"))

class SchemeDetailsView(AdminAndAuthenticatedAccessMixin, View):
    template_name = "admin_app/scheme_details.html"
    template_partial_patient_list = "admin_app/partials/__patients_list.html"
    context_data = {}

    def get(self, request, insurance_number):
        scheme = get_object_or_404(ManagerModels.Scheme, insurance_number=insurance_number)
        self.context_data["scheme"] = scheme

        total_debit = 0
        transactions = ManagerModels.Transaction.objects.filter(scheme=scheme)
        self.context_data["transactions"] = transactions

        for transaction in transactions:
            total_debit = total_debit + transaction.amount_used
        self.context_data["total_debit"] = total_debit

        members = ManagerModels.FamilyMember.objects.filter(scheme=scheme)
        self.context_data["members"] = members

        self.context_data["patients"] = group_patients([m.patient for m in members])

        return render(request, template_name=self.template_name, context=self.context_data)

    def post(self, request, insurance_number):
        scheme = get_object_or_404(ManagerModels.Scheme, insurance_number=insurance_number)
        self.context_data["scheme"] = scheme

        if request.POST.get("suspend_acount"):
            scheme.status = False
            scheme.save()
            return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

        if request.POST.get("activate_acount"):
            scheme.status = True
            scheme.save()
            return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

        patient_data = request.POST.get("search-keyword")
        if not patient_data:
            self.context_data["patients"] = []
            return render(request, template_name=self.template_partial_patient_list, context=self.context_data)

        if len(patient_data.split(" ")) == 2:
            name = patient_data.split(" ")
            patients = ClinicModels.Patients.objects.using("clinic").filter(Q(firstname__icontains=name[0], lastname__icontains=name[1]) | Q(firstname__icontains=name[1], lastname__icontains=name[0]))
        else:
            patients = ClinicModels.Patients.objects.using("clinic").filter(Q(patientid__icontains=patient_data) | Q(patientno__icontains=patient_data) | Q(nationalidno__icontains=patient_data) | Q(referenceno__icontains=patient_data) | Q(firstname__icontains=patient_data) | Q(lastname__icontains=patient_data) | Q(middlename__icontains=patient_data))

        if not patients:
            messages.add_message(request, messages.ERROR, _("No patient records found."))
        
        self.context_data["patients"] = group_patients(patients)
        return render(request, template_name=self.template_partial_patient_list, context=self.context_data)

class EditAccountView(AdminAndAuthenticatedAccessMixin, View):
    template_partial_confirm_form = "admin_app/partials/__edit_scheme.html"
    template_partial_scheme_home = "admin_app/partials/__scheme_home_html"
    context_data = {}

    def get(self, request, insurance_number):
        schemes = ManagerModels.Scheme.objects.filter(insurance_number = insurance_number)
        if not schemes:
            messages.add_message(request, messages.ERROR, _("Record not found."))
            return redirect(reverse("admin_app:schemes"))

        scheme = schemes.first()
        self.context_data["scheme"] = scheme
        
        return render(request, template_name=self.template_partial_confirm_form, context=self.context_data)

    def post(self, request, insurance_number):
        if request.POST.get("account_name"):
            account_name = request.POST.get("account_name")
            
            schemes = ManagerModels.Scheme.objects.filter(insurance_number = insurance_number)
            if not schemes:
                messages.add_message(request, messages.ERROR, _("Record not found."))
                return redirect(reverse("admin_app:schemes"))

            scheme = schemes.first()
            scheme.name = account_name
            scheme.save()
            
            return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))


class SchemeTransactionStatementView(AdminAndAuthenticatedAccessMixin, View):
    template_name = "admin_app/partials/__scheme_home_html"
    context_data = {}

    header_end_row = 15
    excel_cols = [
        {"column": "A", "name": "Date"},
        {"column": "B", "name": "Name"},
        {"column": "C", "name": "Service"},
        {"column": "D", "name": "Debit"},
        {"column": "E", "name": "Credit"},
        {"column": "F", "name": "Cummulative Balanace"},
    ]

    def get(self, request, insurance_number):
        schemes = ManagerModels.Scheme.objects.filter(insurance_number = insurance_number)
        if not schemes:
            messages.add_message(request, messages.ERROR, _("Record not found."))
            return redirect(reverse("admin_app:schemes"))

        scheme = schemes.first()
    
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = scheme.name

        # Load the image file
        img = Image(os.path.join(settings.BASE_DIR , "static/images/statement header.jpg"))
        
        # Resize the image as per the cell size
        img.width = 800
        img.height = 250
        
        # Add the image to the worksheet in cells A1 to F1
        sheet.add_image(img, 'A1')

        # write table column names
        for col in self.excel_cols:
            sheet[f"{col["column"]}{self.header_end_row}"] = col["name"]

        # Create a Font object for bold style
        bold_font = Font(bold=True)
        
        # Apply bold font to the entire row
        for row in sheet.iter_rows(min_row=self.header_end_row, max_row=self.header_end_row):  # Iterate over the first row
            for cell in row:
                cell.font = bold_font
        
        # get scheme transactions in ascending order(old to new)
        transactions = ManagerModels.Transaction.objects.filter(scheme=scheme)
        if not transactions:
            # empty worksheet
            pass
        
        i = 1
        init_account = 0
        for transaction in transactions:

            # get transcation services
            services = ManagerModels.TransactionService.objects.filter(transaction=transaction)
            if services:
                for service in services:
                
                    init_account = init_account - service.service.price
                    sheet[f"{self.excel_cols[0]["column"]}{self.header_end_row + i}"] = transaction.created_on # Date
                    sheet[f"{self.excel_cols[1]["column"]}{self.header_end_row + i}"] = f"{transaction.member.patient.firstname} {transaction.member.patient.lastname}" # Name
                    sheet[f"{self.excel_cols[2]["column"]}{self.header_end_row + i}"] = service.service.name # Service
                    sheet[f"{self.excel_cols[3]["column"]}{self.header_end_row + i}"] = service.service.price # Debit
                    sheet[f"{self.excel_cols[4]["column"]}{self.header_end_row + i}"] = "" # Credit
                    sheet[f"{self.excel_cols[5]["column"]}{self.header_end_row + i}"] = init_account # Cummulative Balanace
                    i = i + 1
            else:
                # credit transaction
                if transaction.reason.lower() == "credit":
                
                    init_account = init_account + transaction.amount_used
                    sheet[f"{self.excel_cols[0]["column"]}{self.header_end_row + i}"] = transaction.created_on # Date
                    sheet[f"{self.excel_cols[1]["column"]}{self.header_end_row + i}"] = transaction.depositor #name
                    sheet[f"{self.excel_cols[2]["column"]}{self.header_end_row + i}"] = "DEPOSIT" # Service
                    sheet[f"{self.excel_cols[3]["column"]}{self.header_end_row + i}"] = "" # Debit
                    sheet[f"{self.excel_cols[4]["column"]}{self.header_end_row + i}"] = transaction.amount_used # Credit
                    sheet[f"{self.excel_cols[5]["column"]}{self.header_end_row + i}"] = init_account # Cummulative Balanace
                    i = i + 1
                else:
                    # not credit
                    pass

            
        # Create a BytesIO object to save the workbook
        output = BytesIO()
        workbook.save(output)
        # Rewind the buffer
        output.seek(0)
        
        # Create a HttpResponse object with the Excel content type
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # Set the file name
        response['Content-Disposition'] = f'attachment; filename={scheme.name}.xlsx'
        return response
    

class CreditAccountView(AdminAndAuthenticatedAccessMixin, View):
    template_partial_confirm_form = "admin_app/partials/__credit_scheme.html"
    template_partial_scheme_home = "admin_app/partials/__scheme_home_html"
    context_data = {}

    def get(self, request, insurance_number):
        schemes = ManagerModels.Scheme.objects.filter(insurance_number = insurance_number)
        if not schemes:
            messages.add_message(request, messages.ERROR, _("Record not found."))
            return redirect(reverse("admin_app:schemes"))

        scheme = schemes.first()
        self.context_data["scheme"] = scheme
        
        return render(request, template_name=self.template_partial_confirm_form, context=self.context_data)

    def post(self, request, insurance_number):
        if request.POST.get("credit"):
            credit_account = request.POST.get("credit_account")

            transac = request.POST.get("transac")
            depositor = request.POST.get("depositor")
            bank_payment = request.POST.get("bank_payment")
            cash_payment = request.POST.get("cash_payment")
            date = request.POST.get("date")

            payment_mode = cash_payment
            if bank_payment:
                payment_mode = payment_mode

            schemes = ManagerModels.Scheme.objects.filter(insurance_number = insurance_number)
            if not schemes:
                messages.add_message(request, messages.ERROR, _("Record not found."))
                return redirect(reverse("admin_app:schemes"))

            scheme = schemes.first()
            scheme.credit = scheme.credit + decimal.Decimal(str(credit_account))
            scheme.save()

            transaction = ManagerModels.Transaction(
                reference_no=transac,
                reason = "Credit",
                depositor = depositor,
                payment_mode = payment_mode,
                scheme=scheme,
                amount_used=decimal.Decimal(str(credit_account)),
                created_on=date,
                completed=True,
                authorised=True,
            )
            transaction.save()

            notify_principle(request, "Scheme Credit", scheme, transaction, f"This email is to inform you that UGX {str(transaction.amount_used)} has been credited to your scheme {scheme.name}.")

            # TODO: make redirect
            self.context_data["scheme"] = scheme
            total_debit = 0
            transactions = ManagerModels.Transaction.objects.filter(scheme=scheme)
            self.context_data["transactions"] = transactions

            for transaction in transactions:
                total_debit = total_debit + transaction.amount_used
            self.context_data["total_debit"] = total_debit

            members = ManagerModels.FamilyMember.objects.filter(scheme=scheme)
            if members:
                members = members[:10]
            self.context_data["members"] = members

            self.context_data["patients"] = group_patients([m.patient for m in members])

            return render(request, template_name=self.template_partial_scheme_home, context=self.context_data)

class SchemmePatientView(AdminAndAuthenticatedAccessMixin, View):
    def get(self, request, insurance_number, patientno):
        pass

    def post(self, request, insurance_number, patientno):
        insurances = ManagerModels.Scheme.objects.filter(insurance_number = insurance_number)
        if not insurances:
            messages.add_message(request, messages.ERROR, _("Scheme record not found"))
            return redirect(reverse("admin_app:schemes"))

        if request.POST.get("assign_patient"):
            if ManagerModels.FamilyMember.objects.filter(patient__patientno=patientno):
                messages.add_message(request, messages.ERROR, _("Patient already has an insurance"))
                return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

            patient = ClinicModels.Patients.objects.using("clinic").filter(patientno=patientno)
            if not patient:
                messages.add_message(request, messages.ERROR, _("Patient record not found"))
                return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

            # save patient to our db
            patient = clone_patient(patient.first())
            if not patient:
                messages.add_message(request, messages.ERROR, _("Incomplete Operation, contact support"))
                return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

            member = ManagerModels.FamilyMember(
                scheme=insurances.first(),
                patient=patient,
                status=False
            )
            member.save()

            messages.add_message(request, messages.SUCCESS, _("Operation Successful. Edit Membership details"))
            return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

        if request.POST.get("remove_patient"):
            memberships = ManagerModels.FamilyMember.objects.filter(patient__patientno=patientno, scheme__insurance_number=insurance_number)
            if not memberships:
                messages.add_message(request, messages.ERROR, _("Invalid data provided"))
                return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

            membership = memberships.first()
            membership.patient.delete()
            membership.delete()

            # log
            log = ManagerModels.Log(
                action=f"Deleted Scheme membership: {membership.scheme_no}. Scheme: {membership.scheme}"
            )
            log.save()

            messages.add_message(request, messages.SUCCESS, _("Operation Successful"))
            return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

        if request.POST.get("activate"):
            memberships = ManagerModels.FamilyMember.objects.filter(patient__patientno=patientno, scheme__insurance_number=insurance_number)
            if not memberships:
                messages.add_message(request, messages.ERROR, _("Invalid data provided"))
                return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

            membership = memberships.first()
            membership.status = True
            membership.save()

            messages.add_message(request, messages.SUCCESS, _("Operation Successful"))
            return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

        if request.POST.get("deactivate"):
            memberships = ManagerModels.FamilyMember.objects.filter(patient__patientno=patientno, scheme__insurance_number=insurance_number)
            if not memberships:
                messages.add_message(request, messages.ERROR, _("Invalid data provided"))
                return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

            membership = memberships.first()
            membership.status = False
            membership.save()

            messages.add_message(request, messages.SUCCESS, _("Operation Successful"))
            return redirect(reverse("admin_app:scheme-details", args=[insurance_number]))

class SchemeCreateView(AdminAndAuthenticatedAccessMixin, View):
    template_partial_home = "admin_app/partials/__create_scheme.html"
    context_data = {}

    def get(self, request):
        form = ManagerForms.SchemeForm()
        self.context_data["form"] = form
        total_credit = 0
        total_debit = 0
        schemes = ManagerModels.Scheme.objects.filter(status=True).order_by("-id")
        transactions = ManagerModels.Transaction.objects.all()[:10]

        for scheme in schemes:
            total_credit = total_credit + scheme.credit

        for transaction in transactions:
            total_debit = total_debit + transaction.amount_used

        self.context_data["total_credit"] = total_credit
        self.context_data["total_debit"] = total_debit
        self.context_data["schemes"] = schemes
        self.context_data["transactions"] = schemes
        self.context_data["beneficiaries"] = ManagerModels.FamilyMember.objects.filter(status=True)

        return render(request, template_name=self.template_partial_home, context=self.context_data)

    def post(self, request):
        form = ManagerForms.SchemeForm(request.POST)
        if form.is_valid:
            form.save()
            messages.add_message(request, messages.SUCCESS, _("Scheme created successfully"))
        else:
            messages.add_message(request, messages.ERROR, _("Unable to create scheme"))
        return redirect("admin_app:schemes")


def group_instances(insurances):
    insurance_groups = []
    for insurance in insurances:
        # get insurance beneficiarys
        beneficiarys = ManagerModels.FamilyMember.objects.filter(family=insurance.family, status=True)

        data = {
            "insurance": insurance,
            "beneficiary_count": len(beneficiarys)
        }
        insurance_groups.append(data)
    return insurance_groups

def group_patients(patients):
    patient_groups = []
    
    for patient in patients:
        scheme = None
        membership_status = None
        relationship = None
        access_type=None
        member = ManagerModels.FamilyMember.objects.filter(patient__patientno=patient.patientno)
        if member:
            scheme = member.first().scheme
            membership_status = member.first().status
            relationship = member.first().relationship
            access_type = member.first().access_type

        data = {
            "patient" : patient,
            "scheme": scheme,
            "membership_status": membership_status,
            'relationship': relationship,
            'access_type': access_type
        }
        patient_groups.append(data)

    return patient_groups

def get_patient(request, patientno, commit=True):
    # check if patient exists in our database
    patients = ManagerModels.Patient.objects.filter(patientno=patientno)
    
    if not patients:
        patients = ClinicModels.Patients.objects.using("clinic").filter(patientno=patientno)
        return clone_patient(patients.first(), commit), False

    if not patients:
        # patient not found in clinicMas database
        messages.add_message(request, messages.ERROR, _("Record not found."))
        return redirect(reverse("admin_app:home"))

    return patients.first(), True

class PatientDetails(AdminAndAuthenticatedAccessMixin, View):
    template_name = "admin_app/patient_details.html"
    template_partial_diodata = "admin_app/partials/__patient_biodata.html"
    context_data = {}

    def get(self, request, patientno):
        patient, is_beneficiary = get_patient(request, patientno, commit=False)

        if not is_beneficiary:
            messages.add_message(self.request, messages.ERROR, _("Patient not a beneficiary of any insurance"))
        else:
            members = ManagerModels.FamilyMember.objects.filter(patient__patientno=patientno)
            if members:
                member = members.first()
                self.context_data["member"] = member
                self.context_data["scheme"] = member.scheme
            
        self.context_data["patient"] = patient
        self.context_data["form"] = ManagerForms.PatientForm(instance=patient)
        
        return render(request, template_name=self.template_name, context=self.context_data)

    def post(self, request, patientno):
        patient, is_beneficiary = get_patient(patientno, commit=False)

        # check is record exists
        instance = ManagerModels.Patient.objects.filter(patientno=request.POST.get("patientno"))
        if instance:
            form = ManagerForms.PatientForm(request.POST, instance=instance.first())
        else:
            form = ManagerForms.PatientForm(request.POST)

        # copy all fields

        if not form.is_valid:
            messages.add_message(request, messages.ERROR, _("Error Processing data."))
        else:
            form.save()
            messages.add_message(request, messages.SUCCESS, _("Operation was successful."))
            
        return redirect(reverse("admin_app:patient-details", args=[patientno]))


class PatientMembershipDetails(AdminAndAuthenticatedAccessMixin, View):
    template_partials_membership= "admin_app/partials/__patient_membership.html"
    context_data = {}
    
    def get(self, request, patientno):
        patient, is_beneficiary = get_patient(request, patientno, commit=False)
        if not is_beneficiary:
            messages.add_message(self.request, messages.ERROR, _("Patient not a benefician of any insurance"))
            
        self.context_data["patient"] = patient
        
        # check if patient is in family
        members = ManagerModels.FamilyMember.objects.filter(patient__patientno=patientno)
        if members:
            member = members.first()
        else:
            member = ManagerModels.FamilyMember(
                patient=patient
            )

        form = ManagerForms.FamilyMemberForm(instance=member)
        self.context_data["form"] = form
        
        return render(request, template_name=self.template_partials_membership, context=self.context_data)
    
    def post(self, request, patientno):
        patient, is_benefician = get_patient(request, patientno)
        
        # check if patient is in family
        members = ManagerModels.FamilyMember.objects.filter(patient__patientno=patientno)
        if members:
            member = members.first()
            form = ManagerForms.FamilyMemberForm(request.POST, instance=member)
        else:
            # adding new member to family
            form = ManagerForms.FamilyMemberForm(request.POST)
            
        if form:
            form.save()
            messages.add_message(request, messages.SUCCESS, _("Details updated/saved successfully"))
        else:
            messages.add_message(request, messages.ERROR, _("Incomplete Operation, contact support"))
            
        return redirect(reverse("admin_app:patient-membership", args=[patientno]))


class PatientPaymentView(AdminAndAuthenticatedAccessMixin, View):
    template_partials_payment= "admin_app/partials/__make_payment.html"
    context_data = {}
    
    def get(self, request, patientno):
        patient, is_beneficiary = get_patient(request, patientno, commit=False)
        if not is_beneficiary:
            messages.add_message(self.request, messages.ERROR, _("Patient not a benefician of any insurance"))
            return redirect(reverse("admin_app:patient-membership", args=[patientno]))
            
        self.context_data["patient"] = patient
        self.context_data["services"] = ManagerModels.Service.objects.all()
        
        return render(request, template_name=self.template_partials_payment, context=self.context_data)
    
    def post(self, request, patientno):
        patient, is_benefician = get_patient(request, patientno)
        
        # check if patient is in family
        members = ManagerModels.FamilyMember.objects.filter(patient__patientno=patientno)
        if members:
            member = members.first()
            form = ManagerForms.FamilyMemberForm(request.POST, instance=member)
        else:
            # adding new member to family
            form = ManagerForms.FamilyMemberForm(request.POST)
            
        if form:
            form.save()
            messages.add_message(request, messages.SUCCESS, _("Details updated/saved successfully"))
        else:
            messages.add_message(request, messages.ERROR, _("Incomplete Operation, contact support"))
            
        return redirect(reverse("admin_app:patient-membership", args=[patientno]))


class ServicesView(AdminAndAuthenticatedAccessMixin, View):
    template_name= "admin_app/services.html"
    context_data = {}
    def get(self, request):
        form = ManagerForms.ServiceForm()
        self.context_data["form"] = form

        self.context_data["services"] = ManagerModels.Service.objects.all()
        
        return render(request, template_name=self.template_name, context=self.context_data)

    def post(self, request):
        if request.FILES.get("excel_file"):
            # TODO: run in background
            excel_file = request.FILES.get("excel_file")

            save_path = os.path.join(settings.MEDIA_ROOT, 'services')
            if not os.path.exists(save_path):
                os.makedirs(save_path)
                
            # Save the uploaded file to disk
            with open(os.path.join(save_path, excel_file.name), 'wb+') as destination:
                for chunk in excel_file.chunks():
                    destination.write(chunk)

            # save contents
            wb = load_workbook(os.path.join(save_path, excel_file.name))
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                code_index = None
                name_index = None
                price_fee_index = None
                for i, value in enumerate(row):
                    if value and "code" in ws.cell(row=1, column=i + 1).value.lower():
                        code_index = i
                    elif value and "name" in ws.cell(row=1, column=i + 1).value.lower():
                        name_index = i
                    elif value and ("price" in ws.cell(row=1, column=i + 1).value.lower() or "fee" in ws.cell(row=1, column=i + 1).value.lower()):
                        price_fee_index = i
                
                if code_index is not None and name_index is not None and price_fee_index is not None:
                    # write to db
                    # data.append({
                    #     'code': row[code_index],
                    #     'name': row[name_index],
                    #     'price_fee': row[price_fee_index]
                    # })
                    ManagerModels.Service.objects.create(
                        code=row[code_index],
                        name=row[name_index],
                        price=row[price_fee_index],
                    )

        else:
            form = ManagerForms.ServiceForm(request.POST)
            if form.is_valid:
                form.save()
                messages.add_message(request, messages.SUCCESS, _("Operation successful"))
            else:
                messages.add_message(request, messages.ERROR, _("Operation not successful"))
        return redirect(reverse("admin_app:services"))


class POSView(AdminAndAuthenticatedAccessMixin, View):
    template_name= "admin_app/pos.html"
    context_data = {}
    def get(self, request):
        self.context_data["services"] = ManagerModels.Service.objects.all()[:10]
        self.context_data["schemes"] = ManagerModels.Scheme.objects.all()[:10]
        return render(request, template_name=self.template_name, context=self.context_data)

    def post(self, request):
        json_data = json.loads(request.body)
        
        member = ManagerModels.FamilyMember.objects.filter(
            patient__patientno=json_data.get("patient").get("patientno"),
            scheme__id=json_data.get("patient").get("scheme")
        )
        
        if not member:
            response_data = {'error': 'Invalid data provided'}
            return JsonResponse(response_data)

        member = member.first()
        scheme = member.scheme
        
        if scheme.credit < json_data.get("total_amount") and not scheme.scheme_type.canExceedCredit:
            response_data = {'error': f'Total amount({json_data.get("total_amount")}) exceed scheme credit({scheme.credit})'}
            return JsonResponse(response_data)

        total_amount = json_data.get("total_amount") * (100 - int(json_data.get("discount"))) / 100

        # create transaction
        transaction = ManagerModels.Transaction.objects.create(
            member = member,
            scheme=scheme,
            amount_used = total_amount,
            reason=f"Payment for Member: {member.patient}",
            completed=True,
            authorised=True,
            patient_type=json_data.get("patient_type")
        )
        
        try:
            for service in json_data.get("services"):
                
                service = ManagerModels.Service.objects.filter(id=service.get("id"), code=service.get("code"))
                if not service:
                    response_data = {'error': 'Invalid data provided'}
                    return JsonResponse(response_data)
                service = service.first()

                ManagerModels.TransactionService.objects.create(
                    service = service,
                    transaction = transaction
                )

                # reduce scheme credit
                scheme.credit = scheme.credit - service.price
                scheme.save()
        except:
            transaction.delete()
            response_data = {'error': 'Error processing data'}
            return JsonResponse(response_data)

        notify_principle(request, transaction.reason, scheme, transaction, f"This email is to inform you that UGX {str(transaction.amount_used)} has been debited from your scheme {scheme.name} for beneficiary {member.patient}.")

        response_data = {'message': 'Data received successfully', "transaction_ref": transaction.reference_no}
        return JsonResponse(response_data, status=200)

class POSSchemePatientsView(AdminAndAuthenticatedAccessMixin, View):
    __patient_partial_template = "admin_app/partials/__pos_patients_list.html"
    context_data = {}
    def get(self, request, insurance_number):
        scheme = get_object_or_404(ManagerModels.Scheme, insurance_number=insurance_number)
        self.context_data["scheme"] = scheme

        members = ManagerModels.FamilyMember.objects.filter(scheme=scheme)
        self.context_data["members"] = members

        self.context_data["patients"] = group_patients([m.patient for m in members])

        return render(request, template_name=self.__patient_partial_template, context=self.context_data)


class POSServicesView(AdminAndAuthenticatedAccessMixin, View):
    __partial_template = "admin_app/partials/__pos_services_list.html"
    context_data = {}
    def post(self, request):
        service_keyword = request.POST.get("service_keyword")
        
        if service_keyword:
            services = ManagerModels.Service.objects.filter(Q(name__icontains=service_keyword) | Q(code__icontains=service_keyword))
            self.context_data["services"] = services
        else:
            self.context_data["services"] = ManagerModels.Service.objects.all()[:10]

        return render(request, template_name=self.__partial_template, context=self.context_data)
    
class POSSchemesView(AdminAndAuthenticatedAccessMixin, View):
    __partial_template = "admin_app/partials/__pos_schemes_list.html"
    context_data = {}
    def post(self, request):
        scheme_keyword = request.POST.get("scheme_keyword")
        
        if scheme_keyword:
            schemes = ManagerModels.Scheme.objects.filter(Q(name__icontains=scheme_keyword) | Q(insurance_number__icontains=scheme_keyword))
            self.context_data["schemes"] = schemes
        else:
            self.context_data["schemes"] = ManagerModels.Scheme.objects.all()

        return render(request, template_name=self.__partial_template, context=self.context_data)

class TransactionsView(AdminAndAuthenticatedAccessMixin, View):
    template_name= "admin_app/transactions.html"
    receipt_template_name = "admin_app/transaction_receipt.html"
    context_data = {}
    def get(self, request):
        ref_no = request.GET.get('ref')

        if ref_no:
            transaction = ManagerModels.Transaction.objects.filter(reference_no=ref_no)
            self.context_data["transaction"] = transaction.first()
            if transaction:
                services = ManagerModels.TransactionService.objects.filter(transaction=transaction.first())
                self.context_data["services"] = services
            
            return render(request, template_name=self.receipt_template_name, context=self.context_data)
        else:
            transactions = ManagerModels.Transaction.objects.all().order_by("-id")
            self.context_data["transactions"] = transactions

        return render(request, template_name=self.template_name, context=self.context_data)


# utils
def generate_insurance_no(family_name):
    insurance_no = f"{random.randint(0,10)}{random.randint(0,10)}{random.randint(0,10)}"
    return insurance_no

def clone_patient(cm_patient, commit=True):
    patient = ManagerModels.Patient(
        patientid = cm_patient.patientid,
        patientno = cm_patient.patientno,
        nationalidno = cm_patient.nationalidno,
        referenceno = cm_patient.referenceno,
        firstname = cm_patient.firstname,
        lastname = cm_patient.lastname,
        middlename = cm_patient.middlename,
        birthdate = cm_patient.birthdate,
        fingerprint = cm_patient.fingerprint,
        birthplace = cm_patient.birthplace,
        address = cm_patient.address,
        occupation = cm_patient.occupation,
        phone = cm_patient.phone,
        email = cm_patient.email,
        joindate = cm_patient.joindate,
        location = cm_patient.location,
        nokname = cm_patient.nokname,
        nokrelationship = cm_patient.nokrelationship,
        nokphone = cm_patient.nokphone,
        defaultbillno = cm_patient.defaultbillno,
        defaultmembercardno = cm_patient.defaultmembercardno,
        defaultmainmembername = cm_patient.defaultmainmembername,
        enforcedefaultbillno = cm_patient.enforcedefaultbillno,
        hidedetails = cm_patient.hidedetails,
        tinnumber = cm_patient.tinnumber,
    )

    if not patient:
        return None

    if commit:
        patient.save()

    return patient

def notify_principle(request, subject, scheme, transaction, message):
    # send email to scheme principle

    # get scheme principle
    principals = ManagerModels.FamilyMember.objects.filter(scheme=scheme, relationship__name__icontains="princi")
    if principals:
        principal = principals.first()

        if not principal.patient.email:
            messages.add_message(request, messages.ERROR, _("Principal email not found."))
            return

        try:
            email_body = render_to_string(
                "email_receipt.html",
                {
                    "principal": principal,
                    "message": message,
                    "transaction": transaction,
                    "scheme": scheme
                },
            )
            email = EmailMessage(
                subject = subject,
                body = email_body,
                from_email = settings.SUPPORT_EMAIL,
                to = [principal.patient.email],
                reply_to = [settings.SUPPORT_EMAIL]
            )
            email.content_subtype = 'html'
            # if image_url:
            #     email.attach_file('image_url')

            email.send(fail_silently=False)
        except:
            messages.add_message(request, messages.ERROR, _("Error sending email receipt."))

        messages.add_message(request, messages.SUCCESS, _(f"Transaction receipt to {principal.patient}"))
    else:
        messages.add_message(request, messages.ERROR, _("Email not sent. not principal found."))

        